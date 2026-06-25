"""Convert a JSON array of objects into a new Excel spreadsheet.

Usage:
    python XLSXgen.py <input.json> [output.xlsx] [--column-width N] [--row-height N]

If no output path is given, the spreadsheet is written next to the
input file with the same name and a .xlsx extension.
"""
import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CHECKBOX_FORMAT = '"☑";;"☐"'
DEFAULT_COLUMN_WIDTH = 20
DEFAULT_ROW_HEIGHT = 15

# Unicode blocks for scripts that read right-to-left.
_RTL_RANGES = (
    (0x0590, 0x05FF),  # Hebrew
    (0x0600, 0x06FF),  # Arabic
    (0x0700, 0x074F),  # Syriac
    (0x0750, 0x077F),  # Arabic Supplement
    (0x08A0, 0x08FF),  # Arabic Extended-A
    (0xFB1D, 0xFB4F),  # Hebrew presentation forms
    (0xFB50, 0xFDFF),  # Arabic presentation forms A
    (0xFE70, 0xFEFF),  # Arabic presentation forms B
)

# A number immediately followed by a short trailing unit word, e.g. "360.00
# kg" or "24 lbs" -- split onto its own line so the unit doesn't get crammed
# next to the number in a narrow numeric column.
_NUMBER_UNIT_PATTERN = re.compile(r"^(-?[\d,]*\.?\d+)\s+([^\s\d][^\s]*)$")

# Two or more consecutive spaces inside a value usually means the source data
# was manually column-aligned with spaces (e.g. a code padded out to line up
# with a label after it). A proportional font breaks that alignment, so such
# cells get a monospace font instead, only when this pattern is detected.
_MULTI_SPACE_PATTERN = re.compile(r" {2,}")
MONOSPACE_FONT = Font(name="Consolas")

# Light/off-white greys that are often OCR artifacts from photo lighting.
# Applied both to header cell fills and data cell backgrounds.
_LIGHT_GREYS = {"D9D9D9", "DDDDDD", "E0E0E0", "E5E5E5", "EBEBEB", "F0F0F0", "F2F2F2", "F5F5F5"}

# Sentinel JSON value marking a cell whose source text couldn't be read at
# all (as opposed to a value that was legibly read as zero). The cell is left
# blank but filled red, so an illegible source field stays visually distinct
# from a genuine zero instead of silently looking the same.
ILLEGIBLE = "<ILLEGIBLE>"
ILLEGIBLE_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

HEADER_FONT = Font(bold=True, color="000000")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Single-line cells vertically center within a row that's been resized taller
# to fit a wrapped neighbor, rather than sticking to the top. Cells that wrap
# *themselves* stay top-aligned instead, so their own first line starts at
# the top of the row and reads down naturally.
DATA_ALIGNMENT = Alignment(vertical="center")
MULTILINE_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
MULTILINE_RTL_ALIGNMENT = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)
RTL_ALIGNMENT = Alignment(vertical="center", horizontal="right", readingOrder=2)
THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
# A heavier border for the table's outer perimeter and the line separating the
# header from the data, so the grid reads with the same visual weight as a
# printed form's frame, instead of every line being the same thin weight.
MEDIUM_SIDE = Side(style="medium")


def _table_border(row: int, col: int, last_col: int, bottom_row: int, table_start: int = 1) -> Border:
    """Border for a cell at (row, col) in a table whose outer edge runs from
    `table_start`/col 1 to `bottom_row`/`last_col`, with a heavier line separating
    the header (rows table_start and table_start+1) from the data (row table_start+2+) too.
    Must be computed once and assigned once -- openpyxl silently drops a *second* `.border =`
    assignment to the same cell, and separately drops border sides set on a
    cell that isn't the top-left "anchor" of a merged range (so this can't
    be applied as a touch-up pass after the fact, and the header/data
    separator is implemented as row table_start+2's top edge rather than table_start+1's bottom
    edge, since header cells are routinely merge non-anchors but data rows never are)."""
    return Border(
        left=MEDIUM_SIDE if col == 1 else THIN_SIDE,
        right=MEDIUM_SIDE if col == last_col else THIN_SIDE,
        top=MEDIUM_SIDE if row in (table_start, table_start + 2) else THIN_SIDE,
        bottom=MEDIUM_SIDE if row == bottom_row else THIN_SIDE,
    )


def _display_width(text: str) -> int:
    """Count display width treating East Asian wide/fullwidth characters
    (Chinese, Japanese, Korean, full-width forms, most emoji) as 2 columns
    wide, since Excel renders them roughly twice as wide as Latin letters."""
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in text)


def _is_rtl(text: str) -> bool:
    """True if `text` is predominantly Hebrew/Arabic/Syriac script, so it
    should be right-aligned with right-to-left reading order."""
    rtl_count = sum(1 for ch in text if any(start <= ord(ch) <= end for start, end in _RTL_RANGES))
    letters = sum(1 for ch in text if ch.isalpha())
    return letters > 0 and rtl_count > letters / 2


def _split_trailing_unit(text: str) -> str:
    """If `text` is a number directly followed by a short unit-like word
    (e.g. "360.00 kg"), move the unit onto its own line. Multi-line values
    and values with more than one space-separated word after the number are
    left alone, so this only fires for the simple "<number> <unit>" shape."""
    if "\n" in text:
        return text
    match = _NUMBER_UNIT_PATTERN.match(text)
    return f"{match.group(1)}\n{match.group(2)}" if match else text


def _wrapped_line_count(text: str, column_width: int) -> int:
    """Estimate how many visual lines `text` will occupy once Excel wraps it
    to fit `column_width`: each explicit newline starts a new line, and any
    segment longer than the column (in display width, so wide scripts like
    CJK count double) wraps into multiple lines on its own."""
    lines = 0
    for segment in text.split("\n"):
        lines += max(1, -(-_display_width(segment) // column_width))  # ceil division
    return lines


def _flatten(record: dict, prefix: str = "") -> tuple[dict, set]:
    """Flatten nested dicts into dot-notated keys. A list of strings is
    treated as a set of flags: each item becomes its own True/False column
    (e.g. Tags: ["vip"] -> Tags.vip = True), tracked in the returned flag-key
    set so missing flags can default to False rather than blank."""
    flat = {}
    flag_keys = set()
    for key, value in record.items():
        flat_key = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            sub_flat, sub_flag_keys = _flatten(value, flat_key)
            flat.update(sub_flat)
            flag_keys.update(sub_flag_keys)
        elif isinstance(value, list) and value and all(isinstance(item, str) for item in value):
            for item in value:
                item_key = f"{flat_key}.{item}"
                flat[item_key] = True
                flag_keys.add(item_key)
        elif isinstance(value, list):
            flat[flat_key] = json.dumps(value, ensure_ascii=False)
        else:
            flat[flat_key] = value
    return flat, flag_keys


def _group_headers(flat_records: list[dict]) -> list[tuple[str | None, list[str]]]:
    """Group dot-notated keys by their top-level prefix, gathering every key
    for a given prefix into one group regardless of where it first appears."""
    group_order: list[str] = []
    group_headers: dict[str, list[str]] = {}
    for record in flat_records:
        for key in record:
            top = key.split(".", 1)[0] if "." in key else key
            if top not in group_headers:
                group_headers[top] = []
                group_order.append(top)
            if key not in group_headers[top]:
                group_headers[top].append(key)

    groups = []
    for top in group_order:
        keys = group_headers[top]
        is_nested = "." in keys[0]
        groups.append((top if is_nested else None, keys))
    return groups


def json_to_xlsx(
    json_path: str,
    xlsx_path: str,
    column_width: float = DEFAULT_COLUMN_WIDTH,
    row_height: float = DEFAULT_ROW_HEIGHT,
    footer: str | None = None,
    column_widths: dict[str, float] | None = None,
    column_alignments: dict[str, str] | None = None,
    column_vertical_alignments: dict[str, str] | None = None,
    column_number_formats: dict[str, str] | None = None,
    column_fills: dict[str, str] | None = None,
    row_fills: dict[int, str] | None = None,
    column_font_colors: dict[str, str] | None = None,
    row_font_colors: dict[int, str] | None = None,
    row_number_formats: dict[int, str] | None = None,
    blank_rows: int = 0,
    header: list[dict[str, str]] | None = None,
) -> None:
    """`column_widths`, `column_alignments`, `column_vertical_alignments`,
    `column_number_formats`, `column_fills`, and `column_font_colors` are
    keyed by leaf header name (e.g. "City" for a nested "Address.City" field)
    and override the `column_width` default / the automatic horizontal/
    vertical alignment / Excel's default "General" number format / the
    (none) fill color / the default black text color for that one column --
    use them to match a physical form's column proportions, per-field text
    alignment (e.g. right-aligning a numeric column, or vertically centering
    a column whose paper counterpart is centered while its neighbors are
    top-aligned), fixed decimal precision (e.g. "0.0000" so a value like 100
    still displays as "100.0000" instead of Excel trimming the trailing
    zeros), shading (e.g. a light gray fill on a "totals" style column), and
    text color (e.g. red for a column of negative adjustments) -- fill/font
    colors are given as hex RGB strings like "D9D9D9" or "FF0000". This only
    colors/shades a cell when a color is actually given for that column/row;
    there's no automatic rule (e.g. nothing colors negative numbers red on
    its own). `row_fills`/`row_font_colors` are keyed by 1-indexed data row
    number (1 = the first record) instead, for shading/coloring whole rows --
    e.g. a summary/total row. `row_number_formats` is the row-keyed
    counterpart to `column_number_formats`, for tables where the metric type
    varies by row instead of by column (e.g. a transposed table with months
    as columns and a "growth rate" row that needs a "%" format while the
    other rows are plain currency) -- it takes precedence over a column-level
    format when both apply to the same cell. `blank_rows` adds that many
    empty rows after the data, still bordered and aligned like the data rows,
    matching a paper form's pre-printed empty rows reserved for entries to be
    filled in later by hand."""
    column_widths = column_widths or {}
    column_alignments = column_alignments or {}
    column_vertical_alignments = column_vertical_alignments or {}
    column_number_formats = column_number_formats or {}
    column_fills = column_fills or {}
    column_font_colors = column_font_colors or {}
    row_font_colors = row_font_colors or {}
    row_fills = row_fills or {}
    row_number_formats = row_number_formats or {}
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # Support both the legacy flat array format and the new structured format.
    if isinstance(raw, list):
        records = raw
        table_row_height_mult = 1
    else:
        records = raw.get("table", {}).get("rows", [])
        if not column_widths:
            column_widths = raw.get("table", {}).get("column_widths") or {}
        if not blank_rows:
            blank_rows = raw.get("table", {}).get("blank_rows", 0)
        if header is None:
            header = raw.get("header")
        # Table-level row_height is a minimum line-count per row (e.g. 2 means rows are at
        # least 2 lines tall). Content can make rows taller; this only sets the floor.
        table_row_height_mult = raw.get("table", {}).get("row_height", 1)

    # Normalise records: cell values may be plain strings or rich cell objects.
    # Strip internal keys before flattening so they don't become columns.
    INTERNAL_KEYS = {"_style", "_value", "_height", "_bg"}

    def _cell_text(cell_val) -> str:
        if isinstance(cell_val, dict):
            return cell_val.get("text", "")
        return cell_val if cell_val is not None else ""

    def _normalised_record(record: dict) -> dict:
        return {k: _cell_text(v) for k, v in record.items() if k not in INTERNAL_KEYS}

    # Drop full_width rows (rendered after blank rows) and duplicate-header subheaders.
    leaf_name_set_early = set(column_widths.keys()) if column_widths else set()

    def _is_dup_header(rec):
        if rec.get("_style") != "subheader":
            return False
        texts = {_cell_text(v) for k, v in rec.items() if k not in INTERNAL_KEYS}
        return bool(texts) and texts <= leaf_name_set_early

    data_records      = [r for r in records
                         if r.get("_style") != "full_width" and not _is_dup_header(r)]
    fullwidth_records = [r for r in records if r.get("_style") == "full_width"]

    # Normalize "CODE  spaces  DESCRIPTOR" alignment per column.
    # Detects cells whose first line is an alphanumeric code followed by whitespace
    # then a descriptor (e.g. "GE0006280   999 購入"). Right-aligns the descriptor
    # to the column width so it sits at the far-right edge of the cell, with all
    # codes left-padded to the same length so the descriptors form a vertical column.
    _CODE_DESCRIPTOR = re.compile(r'^([A-Za-z0-9]+)([ 　]+)(\S.*)$')
    for col_key in ({k for r in data_records for k in r if k not in INTERNAL_KEYS}):
        matches = []
        for rec in data_records:
            cell_val = rec.get(col_key, "")
            text = _cell_text(cell_val)
            first_line = text.split('\n')[0] if text else ""
            m = _CODE_DESCRIPTOR.match(first_line)
            matches.append(m)
        codes = [m.group(1) for m in matches if m]
        if not codes:
            continue
        max_code_len = max(len(c) for c in codes)
        for rec, m in zip(data_records, matches):
            if not m:
                continue
            cell_val = rec.get(col_key, "")
            text = _cell_text(cell_val)
            lines = text.split('\n')
            code, descriptor = m.group(1), m.group(3)
            # Fixed gap: pad code to max_code_len then add 2 spaces before the
            # descriptor. Consistent across sheets regardless of column width.
            lines[0] = code + ' ' * (max_code_len - len(code) + 2) + descriptor
            new_text = '\n'.join(lines)
            if isinstance(cell_val, dict):
                cell_val = dict(cell_val)
                cell_val['text'] = new_text
                rec[col_key] = cell_val
            else:
                rec[col_key] = new_text

    flat_records = []
    flag_keys: set = set()
    for record in data_records:
        flat, record_flag_keys = _flatten(_normalised_record(record))
        flat_records.append(flat)
        flag_keys.update(record_flag_keys)

    groups = _group_headers(flat_records)
    headers = [header for _, group_keys in groups for header in group_keys]
    leaf_names = [header.rsplit(".", 1)[-1] for header in headers]
    col_widths = [column_widths.get(leaf, column_width) for leaf in leaf_names]

    # Ensure each column is wide enough for its actual content (floor = max cell display width).
    for col_index, leaf in enumerate(leaf_names):
        max_content = 0
        for record in flat_records:
            val = record.get(headers[col_index], "")
            if isinstance(val, str):
                max_content = max(max_content, _display_width(val))
        if max_content > col_widths[col_index]:
            col_widths[col_index] = max_content + 2  # +2 for cell padding

    wb = Workbook()
    ws = wb.active

    # Compute grid layout.
    # If a header is present its max_total_span gives us fine-grained grid columns.
    # Each table column is mapped to a proportional number of grid columns so the
    # header proportions and table proportions align on the same Excel columns.
    if header:
        normalised_rows_pre = [
            ([r] if isinstance(r, dict) else r) for r in header
        ]
        max_total_span = max(
            (sum(cd.get("label_span", 1) + cd.get("value_span", 1) for cd in row)
             for row in normalised_rows_pre),
            default=1,
        )

        # Normalise every row so its spans sum to max_total_span.
        # This corrects OCR drift where Claude returns different totals per row.
        for row in normalised_rows_pre:
            row_total = sum(cd.get("label_span", 1) + cd.get("value_span", 1) for cd in row)
            if row_total == max_total_span or len(row) == 0:
                continue
            all_spans = []
            for cd in row:
                all_spans.append(("label_span", cd, cd.get("label_span", 1)))
                all_spans.append(("value_span", cd, cd.get("value_span", 1)))
            # Scale proportionally, greedy-correct the last span.
            remaining = max_total_span
            for i, (key, cd, span) in enumerate(all_spans):
                if i == len(all_spans) - 1:
                    cd[key] = max(0, remaining)
                else:
                    new_span = max(0, round(span / row_total * max_total_span))
                    new_span = min(new_span, remaining - (len(all_spans) - i - 1))
                    cd[key] = new_span
                    remaining -= new_span

        # Majority-vote group boundaries across rows with the same cell count.
        # Rows with the same number of cells should share the same cumulative
        # boundary positions (where one cell ends and the next begins). When one
        # row is the outlier, snap its value_spans to match the majority while
        # keeping label_spans fixed (label widths are small and usually correct).
        from collections import Counter as _Counter
        by_cell_count: dict[int, list[int]] = {}
        for ri, row in enumerate(normalised_rows_pre):
            n = len(row)
            if n > 1:
                by_cell_count.setdefault(n, []).append(ri)
        for n_cells, row_indices in by_cell_count.items():
            if len(row_indices) < 2:
                continue
            def _group_ends(row):
                ends, cum = [], 0
                for cell in row[:-1]:
                    cum += cell.get("label_span", 1) + cell.get("value_span", 1)
                    ends.append(cum)
                return ends
            all_ends = [_group_ends(normalised_rows_pre[ri]) for ri in row_indices]
            majority_ends = [
                _Counter(b[pos] for b in all_ends).most_common(1)[0][0]
                for pos in range(n_cells - 1)
            ]
            for ri, ends in zip(row_indices, all_ends):
                if ends == majority_ends:
                    continue
                row = normalised_rows_pre[ri]
                cum = 0
                for j, cell in enumerate(row):
                    label_s = cell.get("label_span", 1)
                    if j < n_cells - 1:
                        value_s = max(0, majority_ends[j] - cum - label_s)
                    else:
                        value_s = max(0, max_total_span - cum - label_s)
                    cell["label_span"] = label_s
                    cell["value_span"] = value_s
                    cum += label_s + value_s

        n_grid_cols = max(max_total_span, len(headers))
    else:
        n_grid_cols = len(headers)

    # Distribute n_grid_cols among table columns proportionally to their widths.
    total_width = sum(col_widths) or 1
    _rem = n_grid_cols
    table_col_spans: list[int] = []
    for i, w in enumerate(col_widths):
        if i == len(col_widths) - 1:
            table_col_spans.append(max(1, _rem))
        else:
            s = max(1, round(w / total_width * n_grid_cols))
            s = min(s, _rem - (len(col_widths) - i - 1))
            table_col_spans.append(s)
            _rem -= s
    table_col_starts = []
    _cur = 1
    for s in table_col_spans:
        table_col_starts.append(_cur)
        _cur += s
    table_col_ends = [table_col_starts[i] + table_col_spans[i] - 1
                      for i in range(len(headers))]

    # Render the header key-value grid above the main table if present.
    # Each header row is a list of {"label", "value", "label_span", "value_span"} cells.
    # Label and value are rendered as separate adjacent cells.
    header_row_offset = 0
    if header:
        total_cols = n_grid_cols

        def _make_fill(hex_color, default=None):
            color = hex_color if hex_color is not None else default
            if not color:
                return None
            if color.upper().lstrip("#") in _LIGHT_GREYS:
                return None
            return PatternFill(start_color=color.lstrip("#"), end_color=color.lstrip("#"), fill_type="solid")

        def _write_header_cell(text, row, col_start, col_end, bold=False, fill=None, align="left", font_size=2):
            sz = {1: 8, 2: 10, 3: 12}.get(font_size, 10)
            c = ws.cell(row=row, column=col_start, value=text)
            c.font = Font(bold=bold, size=sz)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = THIN_BORDER
            if fill:
                c.fill = fill
            if col_end > col_start:
                ws.merge_cells(start_row=row, end_row=row, start_column=col_start, end_column=col_end)
                for c2 in range(col_start + 1, col_end + 1):
                    cell2 = ws.cell(row=row, column=c2)
                    cell2.border = THIN_BORDER
                    if fill:
                        cell2.fill = fill

        for grid_row in normalised_rows_pre:
            row_height_mult = max((cd.get("height", 1) for cd in grid_row), default=1)

            # Build a flat list of (span, text, bold, fill, align, font_sz) segments.
            segments = []
            for cell_def in grid_row:
                label   = cell_def.get("label", "")
                value   = cell_def.get("value", "")
                ls      = cell_def.get("label_span", 1)
                vs      = cell_def.get("value_span", 1)
                bold    = cell_def.get("bold", True)
                font_sz = cell_def.get("font_size", 2)
                if ls > 0:
                    segments.append((ls, label, bold, None, "left", font_sz))
                if vs > 0:
                    segments.append((vs, value, False, None, "left", font_sz))

            # Greedy last-cell fill: distribute total_cols proportionally, last segment
            # claims all remaining columns so the row always fills exactly total_cols.
            col_cursor     = 1
            remaining_cols = total_cols
            remaining_span = max_total_span
            is_single_full_span = len(segments) == 1
            for i, (span, text, bold, fill, align, font_sz) in enumerate(segments):
                is_last = (i == len(segments) - 1)
                if is_single_full_span:
                    align = "center"
                if is_last:
                    col_count = remaining_cols
                else:
                    col_count = max(1, round(span / remaining_span * remaining_cols))
                    col_count = min(col_count, remaining_cols - (len(segments) - i - 1))
                col_end = col_cursor + col_count - 1
                if col_cursor <= total_cols:
                    _write_header_cell(text, header_row_offset + 1, col_cursor,
                                       min(col_end, total_cols),
                                       bold=bold, fill=fill, align=align, font_size=font_sz)
                col_cursor      = col_end + 1
                remaining_cols -= col_count
                remaining_span -= span

            min_height = 3 if is_single_full_span else 2
            effective_header_height = max(row_height_mult, min_height)
            ws.row_dimensions[header_row_offset + 1].height = DEFAULT_ROW_HEIGHT * effective_header_height
            header_row_offset += 1
        # No blank separator — 調合票 title row (last header row) sits directly above the table.

    R = header_row_offset  # shorthand offset so all table rows shift down cleanly
    leaf_idx = 0
    for top, group_headers in groups:
        gc_start = table_col_starts[leaf_idx]
        if top is None:
            gc_end = table_col_ends[leaf_idx]
            ws.cell(row=R + 1, column=gc_start, value=group_headers[0])
            ws.merge_cells(start_row=R + 1, end_row=R + 2,
                           start_column=gc_start, end_column=gc_end)
            leaf_idx += 1
        else:
            ws.cell(row=R + 1, column=gc_start, value=top)
            for header in group_headers:
                sub_start = table_col_starts[leaf_idx]
                sub_end   = table_col_ends[leaf_idx]
                ws.cell(row=R + 2, column=sub_start,
                        value=header[len(top) + 1:])
                if sub_end > sub_start:
                    ws.merge_cells(start_row=R + 2, end_row=R + 2,
                                   start_column=sub_start, end_column=sub_end)
                leaf_idx += 1
            group_end = table_col_ends[leaf_idx - 1]
            if group_end > gc_start:
                ws.merge_cells(start_row=R + 1, end_row=R + 1,
                               start_column=gc_start, end_column=group_end)

    last_row = R + 2 + len(flat_records) + blank_rows + len(fullwidth_records)
    last_col = n_grid_cols
    bottom_row = last_row + 1 if footer is not None else last_row

    max_data_row_height = row_height
    for record_index, record in enumerate(flat_records, start=1):
        raw_record = data_records[record_index - 1]
        row_index = record_index + R + 2
        line_count = 1
        row_style = raw_record.get("_style", "data")
        # Per-row _height overrides the table default; table default overrides 1.
        row_height_mult = raw_record.get("_height", table_row_height_mult)
        row_fill_color = row_fills.get(record_index)

        for col_index, header in enumerate(headers, start=1):
            leaf = leaf_names[col_index - 1]
            this_column_width = col_widths[col_index - 1]
            gc_start = table_col_starts[col_index - 1]
            gc_end   = table_col_ends[col_index - 1]

            # Pull per-cell style from rich cell object if present.
            raw_cell = raw_record.get(header, raw_record.get(leaf, ""))
            cell_props = raw_cell if isinstance(raw_cell, dict) else {}
            cell_align  = column_alignments.get(leaf)
            _raw_valign = column_vertical_alignments.get(leaf)
            cell_valign = "center" if _raw_valign == "middle" else _raw_valign
            cell_bold   = cell_props.get("bold", False)
            cell_bg     = row_fill_color or column_fills.get(leaf)
            cell_color  = row_font_colors.get(record_index) or column_font_colors.get(leaf)
            cell_wrap   = cell_props.get("wrap", None)

            illegible = False
            if header in flag_keys:
                value = 1 if flat_records[record_index - 1].get(header, False) else 0
                cell = ws.cell(row=row_index, column=gc_start, value=value)
                cell.number_format = CHECKBOX_FORMAT
                cell.alignment = DATA_ALIGNMENT
            else:
                value = flat_records[record_index - 1].get(header)
                illegible = value == ILLEGIBLE
                if illegible:
                    value = None
                elif isinstance(value, str):
                    value = re.sub(r'\n{2,}', '\n', value)
                    value = _split_trailing_unit(value)
                cell = ws.cell(row=row_index, column=gc_start, value=value)

                # Determine wrapping and alignment.
                rtl = isinstance(value, str) and _is_rtl(value)
                should_wrap = cell_wrap if cell_wrap is not None else (
                    isinstance(value, str) and ("\n" in value or _display_width(value) > this_column_width)
                )
                if isinstance(value, str) and _MULTI_SPACE_PATTERN.search(value):
                    cell.font = MONOSPACE_FONT

                # Wrapping cells look better left-aligned; center on a tall wrapped cell
                # pushes each line to the middle which looks wrong for long ingredient text.
                h_align = cell_align or ("right" if rtl else None)
                if should_wrap and h_align == "center":
                    h_align = "left"
                v_align = cell_valign or ("top" if should_wrap else "center")
                cell.alignment = Alignment(
                    horizontal=h_align,
                    vertical=v_align,
                    wrap_text=bool(should_wrap),
                    readingOrder=2 if rtl else 1,
                )
                if should_wrap and isinstance(value, str):
                    line_count = max(line_count, _wrapped_line_count(value, this_column_width))

            number_format_override = row_number_formats.get(record_index) or column_number_formats.get(leaf)
            if number_format_override and header not in flag_keys:
                cell.number_format = number_format_override

            # Apply row-level style first, then per-cell overrides on top.
            # Skip TOTAL_FILL when _bg is a light grey (OCR artifact) — the fill
            # would just add spurious grey to rows that aren't grey in the source.
            font_bold = cell_bold
            if row_style == "total":
                cell.fill = TOTAL_FILL
                font_bold = True
            elif row_style == "subheader":
                cell.fill = SUBHEADER_FILL
                font_bold = True
            if cell_bg:
                cell.fill = PatternFill(start_color=cell_bg, end_color=cell_bg, fill_type="solid")
            if illegible:
                cell.fill = ILLEGIBLE_FILL
            cell.font = Font(
                name=cell.font.name,
                bold=font_bold,
                color=cell_color or "000000",
                size=10,
            )
            cell.border = _table_border(row_index, gc_start, last_col, bottom_row, table_start=R + 1)
            if gc_end > gc_start:
                ws.merge_cells(start_row=row_index, end_row=row_index,
                               start_column=gc_start, end_column=gc_end)
                for c in range(gc_start + 1, gc_end + 1):
                    ws.cell(row=row_index, column=c).border = _table_border(
                        row_index, c, last_col, bottom_row, table_start=R + 1)

        # Row height = max(declared minimum lines, actual content lines) × DEFAULT_ROW_HEIGHT.
        # row_height_mult is the minimum (from per-row _height or table-level row_height);
        # content line count can push it higher but never below the declared minimum.
        effective_height = DEFAULT_ROW_HEIGHT * max(row_height_mult, line_count)
        ws.row_dimensions[row_index].height = effective_height
        max_data_row_height = max(max_data_row_height, effective_height)

    blank_start_row = R + 3 + len(flat_records)
    for row_index in range(blank_start_row, blank_start_row + blank_rows):
        for col_index, header in enumerate(headers, start=1):
            leaf = leaf_names[col_index - 1]
            alignment_override = column_alignments.get(leaf)
            vertical_override = column_vertical_alignments.get(leaf)
            gc_start = table_col_starts[col_index - 1]
            gc_end   = table_col_ends[col_index - 1]
            cell = ws.cell(row=row_index, column=gc_start)
            cell.alignment = (
                Alignment(horizontal=alignment_override, vertical=vertical_override or DATA_ALIGNMENT.vertical)
                if alignment_override or vertical_override
                else DATA_ALIGNMENT
            )
            cell.border = _table_border(row_index, gc_start, last_col, bottom_row, table_start=R + 1)
            if gc_end > gc_start:
                ws.merge_cells(start_row=row_index, end_row=row_index,
                               start_column=gc_start, end_column=gc_end)
                for c in range(gc_start + 1, gc_end + 1):
                    ws.cell(row=row_index, column=c).border = _table_border(
                        row_index, c, last_col, bottom_row, table_start=R + 1)
        ws.row_dimensions[row_index].height = max_data_row_height

    # Full-width rows always go at the very bottom, after all blank rows.
    fw_start_row = blank_start_row + blank_rows
    for fw_index, fw_record in enumerate(fullwidth_records):
        row_index = fw_start_row + fw_index
        text = fw_record.get("_value", "")
        height_mult = fw_record.get("_height", 3)
        # Auto-size: estimate lines from newlines + character wrap, at least height_mult rows.
        total_cell_width = sum(col_widths)
        chars_per_line = max(int(total_cell_width), 20)
        text_lines = sum(
            max(1, (len(ln) + chars_per_line - 1) // chars_per_line)
            for ln in (text or "").split("\n")
        )
        auto_mult = max(height_mult, text_lines)
        cell = ws.cell(row=row_index, column=1, value=text)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top",
                                   wrap_text=bool(text and "\n" in text))
        cell.border = _table_border(row_index, 1, last_col, bottom_row, table_start=R + 1)
        if last_col > 1:
            ws.merge_cells(start_row=row_index, end_row=row_index, start_column=1, end_column=last_col)
            for c in range(2, last_col + 1):
                ws.cell(row=row_index, column=c).border = _table_border(
                    row_index, c, last_col, bottom_row, table_start=R + 1)
        ws.row_dimensions[row_index].height = row_height * auto_mult

    for col_index, header in enumerate(headers, start=1):
        if header in flag_keys:
            column_letter = get_column_letter(table_col_starts[col_index - 1])
            validation = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
            validation.add(f"{column_letter}{R + 3}:{column_letter}{last_row}")
            ws.add_data_validation(validation)

    for row in ws.iter_rows(min_row=R + 1, max_row=R + 2, max_col=n_grid_cols):
        for cell in row:
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = _table_border(cell.row, cell.column, last_col, bottom_row, table_start=R + 1)

    # Each table column's declared width is split evenly across its grid sub-columns.
    for col_index in range(1, len(headers) + 1):
        sub_width = col_widths[col_index - 1] / table_col_spans[col_index - 1]
        for gc in range(table_col_starts[col_index - 1], table_col_ends[col_index - 1] + 1):
            ws.column_dimensions[get_column_letter(gc)].width = sub_width

    if footer is not None:
        footer_row = last_row + 1
        footer_cell = ws.cell(row=footer_row, column=1, value=footer)
        footer_cell.alignment = DATA_ALIGNMENT
        ws.merge_cells(start_row=footer_row, end_row=footer_row, start_column=1, end_column=n_grid_cols)
        for gc in range(1, n_grid_cols + 1):
            ws.cell(row=footer_row, column=gc).border = _table_border(
                footer_row, gc, last_col, bottom_row, table_start=R + 1
            )

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Convert a JSON array of objects into an Excel spreadsheet.")
    parser.add_argument("json_path")
    parser.add_argument("xlsx_path", nargs="?")
    parser.add_argument("--column-width", type=float, default=DEFAULT_COLUMN_WIDTH)
    parser.add_argument("--row-height", type=float, default=DEFAULT_ROW_HEIGHT)
    parser.add_argument("--footer", default=None, help="Text for a merged full-width row at the bottom.")
    parser.add_argument(
        "--column-widths", default=None, help='JSON object mapping header name to width, e.g. \'{"Name": 30}\''
    )
    parser.add_argument(
        "--column-alignments",
        default=None,
        help='JSON object mapping header name to "left"/"center"/"right", e.g. \'{"Price": "right"}\'',
    )
    parser.add_argument(
        "--column-vertical-alignments",
        default=None,
        help='JSON object mapping header name to "top"/"center"/"bottom", e.g. \'{"Price": "center"}\'',
    )
    parser.add_argument(
        "--column-number-formats",
        default=None,
        help='JSON object mapping header name to an Excel number format, e.g. \'{"Price": "0.0000"}\'',
    )
    parser.add_argument(
        "--column-fills",
        default=None,
        help='JSON object mapping header name to a hex fill color, e.g. \'{"Total": "D9D9D9"}\'',
    )
    parser.add_argument(
        "--row-fills",
        default=None,
        help='JSON object mapping 1-indexed data row number to a hex fill color, e.g. \'{"3": "D9D9D9"}\'',
    )
    parser.add_argument(
        "--column-font-colors",
        default=None,
        help='JSON object mapping header name to a hex font color, e.g. \'{"Adjustment": "FF0000"}\'',
    )
    parser.add_argument(
        "--row-font-colors",
        default=None,
        help='JSON object mapping 1-indexed data row number to a hex font color, e.g. \'{"3": "FF0000"}\'',
    )
    parser.add_argument(
        "--row-number-formats",
        default=None,
        help='JSON object mapping 1-indexed data row number to an Excel number format, e.g. \'{"3": "0.0\\"%%\\""}\'',
    )
    parser.add_argument(
        "--blank-rows", type=int, default=0, help="Number of empty bordered rows to add after the data."
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx_path or str(Path(args.json_path).with_suffix(".xlsx"))
    json_to_xlsx(
        args.json_path,
        xlsx_path,
        column_width=args.column_width,
        row_height=args.row_height,
        footer=args.footer,
        column_widths=json.loads(args.column_widths) if args.column_widths else None,
        column_number_formats=(
            json.loads(args.column_number_formats) if args.column_number_formats else None
        ),
        column_vertical_alignments=(
            json.loads(args.column_vertical_alignments) if args.column_vertical_alignments else None
        ),
        column_alignments=json.loads(args.column_alignments) if args.column_alignments else None,
        column_fills=json.loads(args.column_fills) if args.column_fills else None,
        row_fills=(
            {int(k): v for k, v in json.loads(args.row_fills).items()} if args.row_fills else None
        ),
        column_font_colors=json.loads(args.column_font_colors) if args.column_font_colors else None,
        row_font_colors=(
            {int(k): v for k, v in json.loads(args.row_font_colors).items()} if args.row_font_colors else None
        ),
        row_number_formats=(
            {int(k): v for k, v in json.loads(args.row_number_formats).items()}
            if args.row_number_formats
            else None
        ),
        blank_rows=args.blank_rows,
    )


if __name__ == "__main__":
    main()
