import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from json_to_xlsx import json_to_xlsx


class JsonToXlsxTest(unittest.TestCase):
    def test_generates_sheet_from_json(self):
        records = [
            {"Name": "Alice", "Age": "30", "City": "New York"},
            {"Name": "Bob", "Age": "25", "City": "Los Angeles"},
            {"Name": "Charlie", "Age": "35"},
        ]

        with TemporaryDirectory() as tmp_dir:
            json_path = Path(tmp_dir) / "input.json"
            xlsx_path = Path(tmp_dir) / "output.xlsx"
            json_path.write_text(json.dumps(records), encoding="utf-8")

            json_to_xlsx(str(json_path), str(xlsx_path))

            self.assertTrue(xlsx_path.exists())
            wb = load_workbook(xlsx_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            self.assertEqual(rows[0], ("Name", "Age", "City"))
            self.assertEqual(rows[1], ("Alice", "30", "New York"))
            self.assertEqual(rows[2], ("Bob", "25", "Los Angeles"))
            self.assertEqual(rows[3], ("Charlie", "35", ""))


if __name__ == "__main__":
    unittest.main()
