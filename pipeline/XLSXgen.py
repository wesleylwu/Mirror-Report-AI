"""Render an Excel workbook from JSON produced by JSONgen.

Each page in the JSON must have a "template" key containing the declarative
cell-grid schema output by the model.

Usage:
    python XLSXgen.py <json_path> [output.xlsx]
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries


_THIN = Side(style="thin")
_BORDER_ALL = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_BORDER_NONE = Border()

# A4 usable column width in Excel char units (0.5" margins, Calibri 11pt)
_A4_PORTRAIT_COLS  = 100.0
_A4_LANDSCAPE_COLS = 128.0


class _NullCell:
    """Stand-in for a MergedCell position — silently swallows all writes."""
    def __setattr__(self, name, value): pass
    def __getattr__(self, name): return None


def _guard(obj):
    """Replace MergedCell instances with _NullCell, recursively for tuples."""
    if isinstance(obj, MergedCell):
        return _NullCell()
    if isinstance(obj, tuple):
        return tuple(_guard(item) for item in obj)
    return obj


class _SafeWS:
    """Proxy around openpyxl Worksheet that turns MergedCell access into no-ops.

    Claude-generated code often writes to every cell in a range (e.g. to apply
    borders) without checking whether cells are inside a merge region. This proxy
    intercepts those writes and drops them silently so execution continues.
    """
    __slots__ = ("_ws",)

    def __init__(self, ws):
        object.__setattr__(self, "_ws", ws)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_ws"), name)

    def __setattr__(self, name, value):
        setattr(object.__getattribute__(self, "_ws"), name, value)

    def cell(self, row=None, column=None, value=None, **kw):
        ws = object.__getattribute__(self, "_ws")
        c = ws.cell(row=row, column=column)
        if isinstance(c, MergedCell):
            return _NullCell()
        if value is not None:
            c.value = value
        return c

    def __getitem__(self, key):
        return _guard(object.__getattribute__(self, "_ws")[key])

    def __setitem__(self, key, value):
        ws = object.__getattribute__(self, "_ws")
        if not isinstance(ws[key], MergedCell):
            ws[key] = value

    def iter_rows(self, **kw):
        for row in object.__getattribute__(self, "_ws").iter_rows(**kw):
            yield tuple(_NullCell() if isinstance(c, MergedCell) else c for c in row)

    def iter_cols(self, **kw):
        for col in object.__getattribute__(self, "_ws").iter_cols(**kw):
            yield tuple(_NullCell() if isinstance(c, MergedCell) else c for c in col)

    def merge_cells(self, range_string=None, start_row=None, end_row=None,
                    start_column=None, end_column=None):
        # Normalize range_string to row/col coords
        if range_string is not None:
            try:
                sc, sr, ec, er = range_boundaries(str(range_string).upper())
                start_row, end_row, start_column, end_column = sr, er, sc, ec
            except Exception:
                return

        if start_row is None:
            return

        # Single-cell merges are invalid OOXML
        if start_row == end_row and start_column == end_column:
            return

        ws = object.__getattribute__(self, "_ws")

        # Overlapping merges corrupt the file — openpyxl silently writes both,
        # producing invalid XML. Skip any range that touches an existing merge.
        for existing in ws.merged_cells.ranges:
            if not (end_row < existing.min_row or start_row > existing.max_row or
                    end_column < existing.min_col or start_column > existing.max_col):
                return

        try:
            ws.merge_cells(start_row=start_row, end_row=end_row,
                           start_column=start_column, end_column=end_column)
        except Exception:
            pass


def _post_process_sheet(ws) -> None:
    """Hide trailing empty cells, clear spacer borders, fix text wrapping and accidental fills."""
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # Build a set of (row, col) pairs that are anchors of full-width merges
    # (e.g. the title row A1:O1). These carry a value but shouldn't count
    # as real column content — they're just the storage cell for the title.
    half_cols = (max_col or 1) // 2
    full_width_anchors: set = set()
    for mr in ws.merged_cells.ranges:
        if (mr.max_col - mr.min_col) >= half_cols:
            full_width_anchors.add((mr.min_row, mr.min_col))

    # Find the real content boundary.
    # full_width_anchors are excluded from column tracking (so title merges don't
    # inflate eff_max_col) but their ROWS still count toward eff_max_row.
    content_rows: set = set()
    content_cols: set = set()
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if (cell.row, cell.column) in full_width_anchors:
                content_rows.add(cell.row)   # row counts, column does not
                continue
            if cell.value is not None or (cell.fill and cell.fill.fill_type == "solid"):
                content_rows.add(cell.row)
                content_cols.add(cell.column)

    eff_max_row = max(content_rows) if content_rows else max_row
    eff_max_col = max(content_cols) if content_cols else max_col

    # Hide rows/cols entirely outside the content boundary.
    # Also unhide everything inside — Claude sometimes hides blank data rows.
    for r in range(1, eff_max_row + 1):
        ws.row_dimensions[r].hidden = False
    for r in range(eff_max_row + 1, max_row + 1):
        ws.row_dimensions[r].hidden = True
    for c in range(eff_max_col + 1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    # Strip borders from spacer columns: narrow (≤ 6 chars) + no real values
    # + not interior of a merge. Full-width merge anchors don't disqualify a column.
    for c in range(1, eff_max_col + 1):
        if float(ws.column_dimensions[get_column_letter(c)].width or 8.43) > 6.0:
            continue
        is_spacer = True
        for r in range(1, eff_max_row + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                is_spacer = False
                break
            if cell.value is not None and (r, c) not in full_width_anchors:
                is_spacer = False
                break
        if is_spacer:
            for r in range(1, eff_max_row + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.border = _BORDER_NONE

    # Scale columns to fit the page width (columns only — row scaling distorts proportions)
    orientation = (getattr(ws.page_setup, "orientation", None) or "portrait").lower()
    target_cols = _A4_LANDSCAPE_COLS if orientation == "landscape" else _A4_PORTRAIT_COLS
    col_widths = [
        float(ws.column_dimensions[get_column_letter(c)].width or 8.43)
        for c in range(1, eff_max_col + 1)
    ]
    raw_col_total = sum(col_widths) or 1.0
    col_scale = min(1.0, target_cols / raw_col_total)
    if col_scale < 1.0:
        for c, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = max(2.0, w * col_scale)

    _WHITE = {"FFFFFF", "00FFFFFF", "FFFFFFFF"}


def execute_code(code: str, ws) -> None:
    """Execute a Claude-generated build_template(ws) function in place."""
    ns: dict = {}
    exec(compile(code, "<claude_generated>", "exec"), ns)  # noqa: S102
    fn = ns.get("build_template")
    if callable(fn):
        fn(_SafeWS(ws))
    _post_process_sheet(ws)


def render_sheet(data: dict, ws) -> None:
    col_widths  = data.get("col_widths") or []
    rows        = data.get("rows") or []
    orientation = str(data.get("orientation") or "portrait").lower()
    sheet_name  = data.get("sheet_name")
    if sheet_name:
        try:
            ws.title = str(sheet_name)[:31]
        except Exception:
            pass

    # Expand "repeat" shorthand: {"h":N,"repeat":K,"cells":[...]} → K identical rows
    expanded_rows = []
    for row in rows:
        count = int(row.get("repeat") or 1)
        if count > 1:
            base = {k: v for k, v in row.items() if k != "repeat"}
            expanded_rows.extend([base] * count)
        else:
            expanded_rows.append(row)
    rows = expanded_rows

    num_rows = int(data.get("num_rows") or len(rows))
    row_by_r: dict[int, dict] = {}
    for idx, row in enumerate(rows):
        r = row.get("r") or row.get("row") or (idx + 1)
        row_by_r[int(r)] = row

    # Scale columns down to fit A4 (never scale up)
    target_cols = _A4_LANDSCAPE_COLS if orientation == "landscape" else _A4_PORTRAIT_COLS
    raw_col_total = sum(max(1.0, float(w)) for w in col_widths) if col_widths else 1.0
    col_scale = target_cols / raw_col_total

    for i, w in enumerate(col_widths, start=1):
        try:
            ws.column_dimensions[get_column_letter(i)].width = max(4.0, float(w) * col_scale)
        except (ValueError, TypeError):
            ws.column_dimensions[get_column_letter(i)].width = 8.0

    def _render_cell_spec(cell_spec, row_idx):
        try:
            col   = int(cell_spec.get("c") or cell_spec.get("col") or 1)
            span  = max(1, int(cell_spec.get("s") or cell_spec.get("span") or 1))
            rspan = max(1, int(cell_spec.get("rs") or cell_spec.get("rowspan") or 1))
        except (ValueError, TypeError):
            col, span, rspan = 1, 1, 1

        top_left = ws.cell(row=row_idx, column=col)
        if isinstance(top_left, MergedCell):
            return

        value      = cell_spec.get("v") or cell_spec.get("value") or None
        bold       = bool(cell_spec.get("b") or cell_spec.get("bold") or False)
        align      = cell_spec.get("a") or cell_spec.get("align") or "left"
        fill_hex   = cell_spec.get("f") or cell_spec.get("fill")
        x          = cell_spec.get("x") if "x" in cell_spec else cell_spec.get("border")
        end_col    = col + span - 1
        end_row    = row_idx + rspan - 1

        top_left.value     = value
        top_left.font      = Font(bold=bold)
        top_left.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)

        if fill_hex:
            hex_val = str(fill_hex).lstrip("#")
            if len(hex_val) == 6:
                try:
                    top_left.fill = PatternFill(
                        start_color=hex_val, end_color=hex_val, fill_type="solid"
                    )
                except Exception:
                    pass

        # Normalize model output: string "true"/"false" instead of JSON boolean
        if isinstance(x, str):
            if x.lower() == "true":
                x = True
            elif x.lower() in ("false", "none"):
                x = False

        # Determine per-side border flags
        if x is True or x == "all" or x == "tlbr":
            wt = wb = wl = wr = True
        elif isinstance(x, str) and x:
            wt = "t" in x; wb = "b" in x; wl = "l" in x; wr = "r" in x
        else:
            wt = wb = wl = wr = False

        # Stamp borders on every EDGE cell of the span BEFORE merging.
        # openpyxl only lets us style the top-left after merge_cells(), so
        # the right/bottom edges of a multi-cell span must be written now,
        # while all positions are still regular Cell objects.
        for r in range(row_idx, end_row + 1):
            for c in range(col, end_col + 1):
                target = ws.cell(row=r, column=c)
                if isinstance(target, MergedCell):
                    continue
                prev = target.border
                target.border = Border(
                    top    = _THIN if (wt and r == row_idx)  else prev.top,
                    bottom = _THIN if (wb and r == end_row)  else prev.bottom,
                    left   = _THIN if (wl and c == col)      else prev.left,
                    right  = _THIN if (wr and c == end_col)  else prev.right,
                )

        # Merge after borders are stamped
        if span > 1 or rspan > 1:
            try:
                ws.merge_cells(
                    start_row=row_idx, end_row=end_row,
                    start_column=col,  end_column=end_col,
                )
            except Exception:
                pass

    # Render rows and cells
    for row_idx in range(1, num_rows + 1):
        row = row_by_r.get(row_idx, {})
        try:
            height = float(row.get("h") or row.get("height") or 15)
        except (ValueError, TypeError):
            height = 15.0
        ws.row_dimensions[row_idx].height = max(5.0, height)

        for cell_spec in (row.get("cells") or []):
            _render_cell_spec(cell_spec, row_idx)

    # Bump column widths to fit cell text (CJK chars ≈ 2 units, ASCII ≈ 1)
    num_cols = len(col_widths)
    half_cols = num_cols // 2
    for r in range(1, num_rows + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.value or isinstance(cell, MergedCell):
                continue
            text = str(cell.value)
            # Skip cells that are part of a wide merge spanning more than half the columns
            mr_spans = [mr for mr in ws.merged_cells.ranges
                        if mr.min_row <= r <= mr.max_row
                        and mr.min_col <= c <= mr.max_col
                        and (mr.max_col - mr.min_col + 1) > half_cols]
            if mr_spans:
                continue
            needed = sum(2.0 if ord(ch) > 127 else 1.0 for ch in text) * 0.75 + 1
            current = float(ws.column_dimensions[get_column_letter(c)].width or 8.0)
            if needed > current:
                ws.column_dimensions[get_column_letter(c)].width = min(needed, 30.0)

    # Page setup — A4, correct orientation, fit to one page
    ws.page_setup.paperSize   = 9  # A4
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5


def _unique_name(base: str, seen: set) -> str:
    clean = re.sub(r'[:\\/?*\[\]]', "", base)[:30].strip() or "Sheet"
    name, ctr = clean, 1
    while name in seen:
        suffix = f"_{ctr}"
        name = clean[: 30 - len(suffix)] + suffix
        ctr += 1
    seen.add(name)
    return name


def populate_data(ws, data_list: list) -> None:
    """Populate data values into the worksheet."""
    if not data_list:
        return
    from openpyxl.cell.cell import MergedCell
    for item in data_list:
        r = item.get("r") or item.get("row")
        c = item.get("c") or item.get("col")
        v = item.get("v") or item.get("value")
        if r is not None and c is not None and v is not None:
            try:
                cell = ws.cell(row=int(r), column=int(c))
                if not isinstance(cell, MergedCell):
                    cell.value = v
            except Exception as e:
                print(f"Failed to write cell data at row {r}, col {c}: {e}", file=sys.stderr)


def json_to_xlsx(json_path: str, xlsx_path: str) -> None:
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    default_sheet = wb.active
    seen: set = set()

    pages = data.get("pages")
    if pages is None:
        pages = [data]

    for idx, page_data in enumerate(pages):
        if "error" in page_data:
            continue

        code = page_data.get("code", "")
        if code:
            # New code-based approach — extract sheet name from code if present
            m = re.search(r'ws\.title\s*=\s*["\']([^"\']+)["\']', code)
            raw_name = m.group(1) if m else f"Sheet {idx + 1}"
            ws = wb.create_sheet(title=_unique_name(raw_name, seen))
            try:
                execute_code(code, ws)
                populate_data(ws, page_data.get("data"))
            except Exception as e:
                print(f"Code execution failed for page {idx + 1}: {e}", file=sys.stderr)
        else:
            # Legacy JSON template approach
            tmpl = page_data.get("template") or page_data
            raw_name = str(tmpl.get("sheet_name") or f"Sheet {idx + 1}")
            ws = wb.create_sheet(title=_unique_name(raw_name, seen))
            render_sheet(tmpl, ws)
            populate_data(ws, page_data.get("data"))

    if len(wb.worksheets) > 1:
        wb.remove(default_sheet)

    wb.save(xlsx_path)
    print(f"Saved {xlsx_path}", file=sys.stderr)


def main():
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("json_path")
    p.add_argument("xlsx_path", nargs="?")
    args = p.parse_args()
    xlsx_path = args.xlsx_path or str(Path(args.json_path).with_suffix(".xlsx"))
    json_to_xlsx(args.json_path, xlsx_path)


if __name__ == "__main__":
    main()
