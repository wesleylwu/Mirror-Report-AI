import sys; sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
wb = load_workbook('input/IMG_9688 Example.xlsx')
ws = wb.active
print('Merges:')
for m in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
    cell = ws.cell(m.min_row, m.min_col)
    v = repr(str(cell.value or ""))[:30]
    print(f'  R{m.min_row}C{m.min_col}:R{m.max_row}C{m.max_col}  val={v}')
